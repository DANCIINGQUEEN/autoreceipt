from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List
import json
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker, OneCellAnchor
from openpyxl.drawing.picture import PictureFrame, PictureNonVisual
from openpyxl.drawing.xdr import XDRPositiveSize2D
import tempfile
import os

app = FastAPI()

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMPLATE_PATH = os.path.join("assets", "template.xlsx")  # 엑셀 템플릿
STAMP_PATH = os.path.join("assets", "stamp.png")         # 결재 도장 이미지

positions = [
    {"img": ("A4", "B17"), "cells": ["B18", "B19", "B20", "B21", "B22"]},
    {"img": ("C4", "D17"), "cells": ["D18", "D19", "D20", "D21", "D22"]},
    {"img": ("A23", "B36"), "cells": ["B37", "B38", "B39", "B40", "B41"]},
    {"img": ("C23", "D36"), "cells": ["D37", "D38", "D39", "D40", "D41"]},
]

# 🧠 EMU (English Metric Units): 엑셀이 사용하는 내부 단위 (1픽셀 ≈ 9525 EMU) 단위로 거리 계산
def pixels(px): return px * 9525

# 🔥 시트에 결재 도장 이미지 삽입
def insert_stamp_to_sheet(sheet):
    if os.path.exists(STAMP_PATH):
        stamp = ExcelImage(STAMP_PATH)
        stamp.width = 215  # 적당한 사이즈로 조정
        stamp.height = 83
        
        # 픽셀 → EMU 변환
        def px_to_emu(px): return int(px * 9525)

        # 위치 대상: D1 ~ D3
        from_col = 3  # D열
        to_col = 4
        from_row = 0  # 1행 (0-based index)
        to_row = 3

        # D열 너비(px) → 엑셀 열 너비 단위 → 픽셀로 환산
        col_width = sheet.column_dimensions["D"].width or 8.43
        col_width_px = col_width * 7  # 대략 1단위 = 7픽셀

        # 오른쪽 정렬 - 이미지 너비 + 여백 (5px)
        offset_x = px_to_emu(38)

        # D1:D3 행 높이 합산
        total_height_px = 0
        for r in range(1, 4):  # 행 1~3
            height_pt = sheet.row_dimensions[r].height or 15  # 기본 15pt
            px = height_pt * 96 / 72
            total_height_px += px

        # 수직 중앙 정렬
        offset_y = px_to_emu(40)

        # 정확한 위치 앵커
        _from = AnchorMarker(col=from_col, colOff=offset_x, row=from_row, rowOff=offset_y)
        to = AnchorMarker(col=to_col, colOff=0, row=to_row, rowOff=0)

        stamp.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
        
        sheet.add_image(stamp)

@app.post("/export")
async def export_to_excel(
    metadata: str = Form(...),
    files: List[UploadFile] = File(...)
):
    data = json.loads(metadata)
    wb = load_workbook(TEMPLATE_PATH)

    sheet_count = (len(data) + 3) // 4
    file_map = {f.filename: f for f in files}
    temp_paths = []

    for idx, item in enumerate(data):
        sheet = wb[wb.sheetnames[idx // 4]]
        pos = positions[idx % 4]
        image_range = pos["img"]
        cell_keys = pos["cells"]

        # 첫 번째 데이터 삽입 시에만 도장 삽입
        if idx % 4 == 0:
            insert_stamp_to_sheet(sheet)

        # 이미지 삽입
        file = file_map[item["filename"]]
        image_bytes = await file.read()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            tmp.write(image_bytes)
            tmp_path = tmp.name
            temp_paths.append(tmp_path)

        img = ExcelImage(tmp_path)
        img.width = 390
        img.height = 330
        img.anchor = image_range[0]
        sheet.add_image(img)

        # 데이터 삽입
        sheet[cell_keys[0]] = item["date"]
        sheet[cell_keys[1]] = item["name"]
        sheet[cell_keys[2]] = item.get("purpose", "")
        sheet[cell_keys[3]] = item.get("leave_time", "")
        sheet[cell_keys[4]] = float(item["limited_amount"])

       # A3 셀 작성자 이름 동적 삽입 (오직 첫 번째 영수증만 처리)
        if idx == 0:
            sheet["A3"].value = sheet["A3"].value.split("작성자 :")[0] + f"작성자 : {item['name']}"


    # 불필요한 시트 제거 (남는 시트 제거)
    used_sheets = (len(data) + 3) // 4
    while len(wb.sheetnames) > used_sheets:
        del wb[wb.sheetnames[-1]]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    for path in temp_paths:
        try:
            os.remove(path)
        except:
            pass

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=receipt_final.xlsx"}
    )
